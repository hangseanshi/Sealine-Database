#!/usr/bin/env python3
"""
agent.py — Core Sealine Claude agent: ClaudeChat class and all tool implementations.

Provides multi-turn conversation, auto-loads Markdown (.md) files as cached
context, and exposes tools: execute_sql, generate_route_map, create_excel,
send_email. Imported by api.py (REST/web interface).

Requirements:
    pip install anthropic httpx pyodbc python-dotenv openpyxl geopy
    pip install google-auth google-auth-oauthlib google-api-python-client
    Set ANTHROPIC_API_KEY environment variable or create .env file.
"""

import os
import sys
import glob
import json
import base64
import tempfile
import argparse
import textwrap
import httpx
import anthropic
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv

# Load environment variables from .env in the same directory as this script
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(_SCRIPT_DIR, ".env"))

try:
    import pyodbc
    PYODBC_AVAILABLE = True
except ImportError:
    PYODBC_AVAILABLE = False

# ── Geocoding helper ───────────────────────────────────────────────────────────
import math, time
from collections import defaultdict

_geocode_cache: dict[str, tuple[float, float] | None] = {}

def _haversine_km(lat1, lng1, lat2, lng2) -> float:
    """Return great-circle distance in km between two lat/lng points."""
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

def _geocode(location_name: str) -> tuple[float, float] | None:
    """
    Geocode a location name via Nominatim with fallback variants.
    Returns (lat, lng) or None.
    """
    if location_name in _geocode_cache:
        return _geocode_cache[location_name]

    try:
        from geopy.geocoders import Nominatim
        geo = Nominatim(user_agent="sealine-route-map/1.0")

        # Build a list of candidate queries to try in order
        candidates = [location_name]
        # If name has a " - " separator, also try the part after it (e.g. "AMBARLI")
        if " - " in location_name:
            suffix = location_name.split(" - ")[-1].strip()
            candidates.append(suffix)
            # Also try stripping common generic words from the suffix
            for drop in ("TERMINAL", "PORT", "CONTAINER", "DEPOT"):
                if suffix.upper().startswith(drop):
                    candidates.append(suffix[len(drop):].strip())
        # Try title-cased version (Nominatim is sometimes case-sensitive)
        candidates.append(location_name.title())

        val = None
        for query in candidates:
            if not query:
                continue
            time.sleep(1.1)
            result = geo.geocode(query, timeout=10)
            if result:
                val = (result.latitude, result.longitude)
                break
    except Exception:
        val = None

    _geocode_cache[location_name] = val
    return val

def build_coord_fix_map(rows: list) -> dict[str, tuple[float, float]]:
    """
    Detect coordinate collisions: when two distinct location names share the
    same (lat, lng) in the DB, those coords are leaked/wrong. Geocode only the
    conflicting names to get correct coordinates. Unique DB coords are trusted.

    rows: list of (cnum, seq, loc, lat, lng, date, actual, desc, ...)
    Returns: {location_name -> (corrected_lat, corrected_lng)}
    """
    # Group location names by rounded coordinate key
    coord_to_names: dict[tuple, set] = defaultdict(set)
    for r in rows:
        loc, lat, lng = r[2], r[3], r[4]
        if loc and lat is not None and lng is not None:
            coord_to_names[(round(lat, 3), round(lng, 3))].add(loc)

    # Coords shared by more than one distinct name are suspect
    suspect_names: set[str] = set()
    for names in coord_to_names.values():
        if len(names) > 1:
            suspect_names.update(names)

    # Geocode only the suspect names; compare geocoded result to DB coords
    fix_map: dict[str, tuple[float, float]] = {}
    for r in rows:
        loc, lat, lng = r[2], r[3], r[4]
        if loc not in suspect_names or loc in fix_map:
            continue
        geo = _geocode(loc)
        if geo is None:
            continue
        if lat is None or lng is None or _haversine_km(lat, lng, geo[0], geo[1]) > 200.0:
            fix_map[loc] = geo  # DB coords are wrong for this name

    return fix_map

# ── ANSI colour helpers ────────────────────────────────────────────────────────
RESET   = "\033[0m"
BOLD    = "\033[1m"
DIM     = "\033[2m"
CYAN    = "\033[36m"
GREEN   = "\033[32m"
YELLOW  = "\033[33m"
RED     = "\033[31m"
MAGENTA = "\033[35m"
BLUE    = "\033[34m"

def supports_color() -> bool:
    return hasattr(sys.stdout, "isatty") and sys.stdout.isatty()

def c(text: str, *codes: str) -> str:
    if not supports_color():
        return text
    return "".join(codes) + text + RESET


# ── DB connection ──────────────────────────────────────────────────────────────
DB_CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=ushou102-exap1;"
    "DATABASE=searates;"
    "UID=sean;"
    "PWD=4peiling;"
)
MAX_ROWS = 500  # cap rows returned to Claude to avoid token explosion


def run_sql(query: str) -> str:
    """Execute a SQL query and return results as formatted text."""
    q = query.strip()
    # Safety: only allow read operations
    first_word = q.split()[0].upper() if q.split() else ""
    if first_word not in ("SELECT", "WITH", "EXEC", "EXECUTE"):
        return "ERROR: Only SELECT / WITH / EXEC queries are permitted."
    try:
        conn = pyodbc.connect(DB_CONN_STR, timeout=30)
        cursor = conn.cursor()
        cursor.execute(q)

        if cursor.description is None:
            conn.close()
            return "Query executed successfully (no rows returned)."

        cols = [d[0] for d in cursor.description]
        rows = cursor.fetchmany(MAX_ROWS + 1)
        truncated = len(rows) > MAX_ROWS
        rows = rows[:MAX_ROWS]
        conn.close()

        if not rows:
            return f"Columns: {', '.join(cols)}\n(0 rows)"

        # Build a simple pipe-delimited table
        col_widths = [len(col) for col in cols]
        str_rows = []
        for row in rows:
            str_row = [str(v) if v is not None else "NULL" for v in row]
            for i, val in enumerate(str_row):
                col_widths[i] = max(col_widths[i], min(len(val), 50))
            str_rows.append(str_row)

        def fmt_row(values):
            return "  ".join(v[:50].ljust(col_widths[i]) for i, v in enumerate(values))

        header = fmt_row(cols)
        separator = "  ".join("-" * w for w in col_widths)
        lines = [header, separator] + [fmt_row(r) for r in str_rows]
        if truncated:
            lines.append(f"\n(Showing first {MAX_ROWS} rows — results truncated)")
        else:
            lines.append(f"\n({len(rows)} row{'s' if len(rows) != 1 else ''})")

        return "\n".join(lines)

    except Exception as e:
        return f"SQL ERROR: {e}"


# ── Gmail / Excel setup ────────────────────────────────────────────────────────
OPENEXXON_DIR = os.path.join(os.path.dirname(_SCRIPT_DIR), "OpenExxon")

# Gmail credentials resolved from OpenExxon .env
_gmail_creds_path  = os.path.join(OPENEXXON_DIR, os.getenv("GMAIL_CREDENTIALS_PATH", "credentials.json"))
_gmail_token_path  = os.path.join(OPENEXXON_DIR, os.getenv("GMAIL_TOKEN_PATH",        "token.json"))
_gmail_from_addr   = os.getenv("GMAIL_ADDRESS", "")

_gmail_service = None   # initialized lazily on first email send

GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]


def _init_gmail():
    """Initialize Gmail API service (lazy, called only when needed)."""
    global _gmail_service
    if _gmail_service is not None:
        return True, None
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build

        creds = None
        if os.path.exists(_gmail_token_path):
            creds = Credentials.from_authorized_user_file(_gmail_token_path, GMAIL_SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        if not creds or not creds.valid:
            return False, "Gmail token invalid or missing. Re-authorize via OpenExxon."
        _gmail_service = build("gmail", "v1", credentials=creds)
        return True, None
    except Exception as e:
        return False, f"Gmail init failed: {e}"


def create_excel_file(title: str, columns: list, rows: list, filename: str) -> tuple[str, str | None]:
    """
    Create an Excel file from column headers + rows.
    Returns (filepath, error_message). filepath is None on error.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name max 31 chars

        # Header row — blue background, white bold text
        HEADER_FILL  = PatternFill("solid", fgColor="1F4788")
        HEADER_FONT  = Font(bold=True, color="FFFFFF")
        HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = HEADER_ALIGN

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approx)
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        ws.freeze_panes = "A2"  # freeze header row

        # Save to project directory so it can be served via /files/
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        filepath = os.path.join(_SCRIPT_DIR, filename)
        wb.save(filepath)
        return filepath, None

    except ImportError:
        return None, "openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return None, f"Excel creation failed: {e}"


def generate_route_map(track_number: str) -> tuple[str, str | None]:
    """
    Query DB for a tracking number's container events and generate a Leaflet.js HTML map.
    Returns (url_path, error_message). url_path is a /files/ URL on success.
    """
    if not PYODBC_AVAILABLE:
        return None, "pyodbc not available — cannot query database."
    try:
        conn = pyodbc.connect(DB_CONN_STR, timeout=15)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                e.Container_NUMBER,
                TRY_CAST(e.Order_Id AS INT)  AS Seq,
                COALESCE(f.name, l.Name)     AS LocationName,
                COALESCE(
                    TRY_CAST(f.Lat AS FLOAT), TRY_CAST(l.Lat AS FLOAT)
                )                            AS Lat,
                COALESCE(
                    TRY_CAST(f.Lng AS FLOAT), TRY_CAST(l.Lng AS FLOAT)
                )                            AS Lng,
                e.Date,
                e.Actual,
                e.Description,
                h.Sealine_Code,
                h.Status
            FROM Sealine_Container_Event e
            INNER JOIN Sealine_Header h
                ON e.TrackNumber = h.TrackNumber AND h.DeletedDt IS NULL
            LEFT JOIN Sealine_Facilities f
                ON e.TrackNumber = f.TrackNumber AND e.Facility = f.Id AND f.DeletedDt IS NULL
            LEFT JOIN Sealine_Locations l
                ON e.TrackNumber = l.TrackNumber AND e.Location = l.Id AND l.DeletedDt IS NULL
            WHERE e.TrackNumber = ?
              AND e.DeletedDt IS NULL
            ORDER BY e.Container_NUMBER, TRY_CAST(e.Order_Id AS INT)
        """, (track_number,))
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            return None, f"No events found for tracking number: {track_number}"

        # Build per-container event lists (only those with coordinates)
        import json as _json
        containers: dict[str, list] = {}
        all_coords = []
        sealine_code = rows[0][8] if rows else ""
        status       = rows[0][9] if rows else ""

        # Build coord fix map: only geocode names with duplicate/leaked coords
        fix_map = build_coord_fix_map(rows)

        seen: set = set()
        for r in rows:
            cnum, seq, loc, lat, lng, date, actual, desc, _, _ = r
            # Apply geocoded correction if this location's DB coords are wrong
            if loc in fix_map:
                lat, lng = fix_map[loc]
            if lat is None or lng is None:
                continue
            # Deduplicate by (cnum, seq, rounded coords)
            key = (cnum, seq, round(lat, 4), round(lng, 4))
            if key in seen:
                continue
            seen.add(key)
            if cnum not in containers:
                containers[cnum] = []
            dt_str = date.strftime("%Y-%m-%d") if date else ""
            containers[cnum].append({
                "seq": seq or 0, "loc": loc or "Unknown",
                "lat": lat, "lng": lng,
                "date": dt_str,
                "actual": bool(actual),
                "desc": desc or ""
            })
            all_coords.append([lat, lng])

        if not all_coords:
            return None, "No georeferenced events found for this tracking number."

        # Map centre
        avg_lat = sum(c[0] for c in all_coords) / len(all_coords)
        avg_lng = sum(c[1] for c in all_coords) / len(all_coords)

        containers_json = _json.dumps(containers)
        track_json      = _json.dumps(track_number)
        sealine_json    = _json.dumps(sealine_code)
        status_json     = _json.dumps(status)

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Route: {track_number}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<style>
  html,body{{margin:0;padding:0;height:100%;font-family:Arial,sans-serif;background:#0d1f33;color:#cde;}}
  #header{{padding:8px 16px;background:#0f2744;border-bottom:2px solid #2e5f8a;display:flex;align-items:center;gap:14px;flex-wrap:wrap;}}
  #header h1{{font-size:15px;color:#7eb8e6;margin:0;white-space:nowrap;}}
  #header span{{font-size:11px;color:#aac8e0;}}
  #map{{height:calc(100vh - 44px);}}
  .popup-box{{font-size:12px;min-width:220px;line-height:1.6;}}
  .popup-box .popup-title{{font-weight:bold;font-size:13px;color:#1a3a6c;border-bottom:1px solid #ccc;padding-bottom:4px;margin-bottom:6px;}}
  .popup-box .popup-row{{display:flex;justify-content:space-between;gap:8px;}}
  .popup-box .popup-label{{color:#555;font-weight:bold;}}
  .popup-box .popup-val{{color:#222;text-align:right;}}
  .actual{{color:#22aa55;font-weight:bold;}}
  .estimated{{color:#cc8800;font-weight:bold;}}
</style>
</head>
<body>
<div id="header">
  <h1>&#128230; {track_number}</h1>
  <span id="hdr-info"></span>
</div>
<div id="map"></div>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet-polylinedecorator@1.6.0/dist/leaflet.polylineDecorator.js"></script>
<script>
const TRACK      = {track_json};
const SEALINE    = {sealine_json};
const STATUS     = {status_json};
const CONTAINERS = {containers_json};

document.getElementById('hdr-info').textContent =
  SEALINE + ' \u2022 ' + STATUS + ' \u2022 ' + Object.keys(CONTAINERS).length + ' containers';

const map = L.map('map',{{center:[{avg_lat},{avg_lng}],zoom:3,preferCanvas:false}});
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/dark_all/{{z}}/{{x}}/{{y}}{{r}}.png',
  {{attribution:'&copy; OpenStreetMap &copy; CARTO',maxZoom:18}}).addTo(map);

const COLORS = ['#4ec9ff','#ff6b6b','#69ff47','#ffe119','#f58231',
                '#b96ff5','#42d4f4','#f032e6','#bfef45','#fabed4'];
let ci = 0;

Object.entries(CONTAINERS).forEach(([cnum, events]) => {{
  const col = COLORS[ci++ % COLORS.length];
  const pts = events.map(e => [e.lat, e.lng]);

  // Draw route line + arrowhead decorators
  if (pts.length > 1) {{
    // Full polyline for visual continuity
    L.polyline(pts, {{color:col, weight:3, opacity:0.8}}).addTo(map);

    // Add an arrow to every individual segment
    for (let i = 0; i < pts.length - 1; i++) {{
      const seg = L.polyline([pts[i], pts[i+1]], {{color:col, weight:3, opacity:0}}).addTo(map);
      L.polylineDecorator(seg, {{
        patterns: [
          {{
            offset: '50%',
            repeat: 0,
            symbol: L.Symbol.arrowHead({{
              pixelSize: 12,
              polygon: false,
              pathOptions: {{color: col, fillOpacity:1, weight:2}}
            }})
          }}
        ]
      }}).addTo(map);
    }}
  }}

  // Draw markers
  events.forEach((e,i) => {{
    const isFirst = i===0, isLast = i===events.length-1;
    const r = isFirst ? 10 : isLast ? 10 : 6;
    const marker = L.circleMarker([e.lat,e.lng],{{
      radius:r,
      fillColor: isFirst ? '#22dd66' : isLast ? '#ff4444' : col,
      color:'#fff', weight:2,
      opacity:1, fillOpacity: isFirst||isLast ? 1 : 0.85
    }}).addTo(map);

    const badge = e.actual
      ? '<span class="actual">&#10003; Actual</span>'
      : '<span class="estimated">&#9711; Estimated</span>';

    marker.bindPopup(`
      <div class="popup-box">
        <div class="popup-title">&#128230; ${{TRACK}}</div>
        <div class="popup-row"><span class="popup-label">Container</span><span class="popup-val">${{cnum}}</span></div>
        <div class="popup-row"><span class="popup-label">Order ID</span><span class="popup-val">${{e.seq}}</span></div>
        <div class="popup-row"><span class="popup-label">Location</span><span class="popup-val">${{e.loc}}</span></div>
        <div class="popup-row"><span class="popup-label">Date</span><span class="popup-val">${{e.date}} ${{badge}}</span></div>
        <div class="popup-row"><span class="popup-label">Event</span><span class="popup-val">${{e.desc}}</span></div>
      </div>`);
  }});
}});

// Legend
const legend = L.control({{position:'bottomleft'}});
legend.onAdd = () => {{
  const d = L.DomUtil.create('div','');
  d.style.cssText='background:#0f2744;padding:8px 12px;border-radius:6px;border:1px solid #2e5f8a;font-size:11px;color:#cde;line-height:1.8;';
  d.innerHTML=`<b style="color:#7eb8e6">Legend</b><br>
    <span style="color:#22dd66">&#9679;</span> Origin &nbsp;
    <span style="color:#ff4444">&#9679;</span> Destination<br>
    &#10230; Arrow = travel direction`;
  return d;
}};
legend.addTo(map);
</script>
</body>
</html>"""

        filename = f"route_{track_number}.html"
        filepath = os.path.join(_SCRIPT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)
        return f"/files/{filename}", None

    except Exception as e:
        return None, f"Map generation failed: {e}"


def send_email_with_attachment(
    to: str,
    subject: str,
    body_html: str,
    attachment_path: str | None = None,
) -> str:
    """Send email via Gmail API, optionally with a file attachment. Returns status string."""
    ok, err = _init_gmail()
    if not ok:
        return f"EMAIL ERROR: {err}"

    try:
        from googleapiclient.errors import HttpError

        # Build MIME message
        if attachment_path and os.path.exists(attachment_path):
            msg = MIMEMultipart()
            msg["to"]      = to
            msg["from"]    = _gmail_from_addr
            msg["subject"] = subject
            msg.attach(MIMEText(body_html, "html"))

            # Attach file
            with open(attachment_path, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            att_name = os.path.basename(attachment_path)
            part.add_header("Content-Disposition", f'attachment; filename="{att_name}"')
            msg.attach(part)
        else:
            msg = MIMEMultipart()
            msg["to"]      = to
            msg["from"]    = _gmail_from_addr
            msg["subject"] = subject
            msg.attach(MIMEText(body_html, "html"))

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
        result = _gmail_service.users().messages().send(
            userId="me", body={"raw": raw}
        ).execute()
        return f"Email sent to {to} (id={result.get('id','')})"

    except Exception as e:
        return f"EMAIL ERROR: {e}"


def generate_location_map(
    title: str,
    markers: list,
    filename: str,
) -> tuple[str, str | None]:
    """
    Generate a Leaflet.js HTML map with a set of marker pins.
    markers: list of dicts with keys: lat, lng, label, popup (optional), color (optional).
    Returns (url_path, error_message).
    """
    try:
        import json as _json

        # Build JS markers array
        marker_js_parts = []
        for m in markers:
            lat   = m.get("lat")
            lng   = m.get("lng")
            label = m.get("label", "")
            popup = m.get("popup", label)
            color = m.get("color", "#4e9af1")
            if lat is None or lng is None:
                continue
            popup_escaped = popup.replace("'", "\\'").replace("\n", "<br>")
            marker_js_parts.append(
                f"addMarker({lat}, {lng}, '{color}', '{popup_escaped}');"
            )

        if not marker_js_parts:
            return None, "No markers with valid coordinates to plot."

        markers_js = "\n    ".join(marker_js_parts)

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #0d1b2a; font-family: 'Segoe UI', sans-serif; }}
  #title-bar {{
    background: #132032;
    color: #e0eaf4;
    font-size: 14px;
    font-weight: 600;
    padding: 8px 14px;
    border-bottom: 1px solid #1e3a5a;
    letter-spacing: 0.03em;
  }}
  #map {{ width: 100%; height: calc(100vh - 36px); }}
  .leaflet-popup-content-wrapper {{
    background: #132032; color: #c8d8e8;
    border: 1px solid #1e3a5a; border-radius: 8px;
  }}
  .leaflet-popup-tip {{ background: #132032; }}
</style>
</head>
<body>
<div id="title-bar">{title}</div>
<div id="map"></div>
<script>
  const map = L.map('map', {{ zoomControl: true }}).setView([20, 0], 2);
  L.tileLayer('https://{{s}}.basemaps.cartocdn.com/dark_all/{{z}}/{{x}}/{{y}}{{r}}.png', {{
    attribution: '&copy; OpenStreetMap &amp; CARTO', maxZoom: 18
  }}).addTo(map);

  function pinIcon(color) {{
    return L.divIcon({{
      className: '',
      html: `<svg width="22" height="30" viewBox="0 0 22 30" xmlns="http://www.w3.org/2000/svg">
        <ellipse cx="11" cy="28" rx="5" ry="2" fill="rgba(0,0,0,0.3)"/>
        <path d="M11 0 C5 0 0 5 0 11 C0 19 11 30 11 30 C11 30 22 19 22 11 C22 5 17 0 11 0Z"
              fill="${{color}}" stroke="#fff" stroke-width="1.5"/>
        <circle cx="11" cy="11" r="4" fill="#fff" opacity="0.85"/>
      </svg>`,
      iconSize: [22, 30],
      iconAnchor: [11, 30],
      popupAnchor: [0, -30]
    }});
  }}

  function addMarker(lat, lng, color, popup) {{
    L.marker([lat, lng], {{ icon: pinIcon(color) }})
      .bindPopup(popup, {{ maxWidth: 320 }})
      .addTo(map);
  }}

  {markers_js}
</script>
</body>
</html>"""

        if not filename.endswith(".html"):
            filename += ".html"
        filepath = os.path.join(_SCRIPT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

        return f"/files/{filename}", None

    except Exception as e:
        return None, f"Location map generation failed: {e}"


def generate_chart(
    title: str,
    chart_type: str,
    labels: list,
    datasets: list,
    filename: str,
    x_label: str = "",
    y_label: str = "",
) -> tuple[str, str | None]:
    """
    Generate a self-contained Chart.js HTML chart file.
    Returns (url_path, error_message). url_path is a /files/ URL on success.
    """
    try:
        import json as _json
        import random

        # Palette — enough colours for up to 20 series/slices
        PALETTE = [
            "#4e9af1", "#f16a4e", "#4ef1a0", "#f1d44e", "#a04ef1",
            "#4ef1e8", "#f14e9a", "#9af14e", "#f1984e", "#4e6af1",
            "#e84ef1", "#4ef168", "#f1c84e", "#4eaef1", "#f14e4e",
            "#b4f14e", "#4ef1d4", "#f18a4e", "#7a4ef1", "#f1f14e",
        ]

        def _colors(n):
            return [PALETTE[i % len(PALETTE)] for i in range(n)]

        # Build dataset config
        ds_configs = []
        is_multi_color = chart_type in ("pie", "doughnut")
        for i, ds in enumerate(datasets):
            cfg = {
                "label": ds.get("label", f"Series {i+1}"),
                "data": ds.get("data", []),
            }
            if is_multi_color:
                cfg["backgroundColor"] = _colors(len(ds.get("data", [])))
                cfg["borderColor"] = "#1a2636"
                cfg["borderWidth"] = 2
            else:
                color = PALETTE[i % len(PALETTE)]
                cfg["backgroundColor"] = color + "aa"  # slight transparency
                cfg["borderColor"] = color
                cfg["borderWidth"] = 2
                if chart_type == "line":
                    cfg["fill"] = False
                    cfg["tension"] = 0.3
                    cfg["pointRadius"] = 4
            ds_configs.append(cfg)

        # Axis options (not applicable for pie/doughnut)
        scales_js = ""
        if chart_type not in ("pie", "doughnut"):
            scales_js = _json.dumps({
                "x": {
                    "title": {"display": bool(x_label), "text": x_label, "color": "#a0b4c8"},
                    "ticks": {"color": "#a0b4c8"},
                    "grid":  {"color": "#2a3f55"},
                },
                "y": {
                    "title": {"display": bool(y_label), "text": y_label, "color": "#a0b4c8"},
                    "ticks": {"color": "#a0b4c8"},
                    "grid":  {"color": "#2a3f55"},
                    "beginAtZero": True,
                },
            })

        chart_cfg = {
            "type": "bar" if chart_type == "horizontalBar" else chart_type,
            "data": {
                "labels": labels,
                "datasets": ds_configs,
            },
            "options": {
                "indexAxis": "y" if chart_type == "horizontalBar" else "x",
                "responsive": True,
                "maintainAspectRatio": False,
                "plugins": {
                    "legend": {
                        "labels": {"color": "#c8d8e8"},
                    },
                    "title": {
                        "display": True,
                        "text": title,
                        "color": "#e0eaf4",
                        "font": {"size": 16},
                    },
                },
            },
        }
        if scales_js:
            chart_cfg["options"]["scales"] = _json.loads(scales_js)

        cfg_json = _json.dumps(chart_cfg, indent=2)

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    background: #0d1b2a;
    color: #c8d8e8;
    font-family: 'Segoe UI', sans-serif;
    display: flex;
    flex-direction: column;
    align-items: center;
    padding: 16px;
    min-height: 100vh;
  }}
  .chart-container {{
    width: 100%;
    max-width: 900px;
    height: 420px;
    background: #132032;
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.4);
  }}
</style>
</head>
<body>
<div class="chart-container">
  <canvas id="chart"></canvas>
</div>
<script>
const cfg = {cfg_json};
new Chart(document.getElementById('chart'), cfg);
</script>
</body>
</html>"""

        if not filename.endswith(".html"):
            filename += ".html"
        filepath = os.path.join(_SCRIPT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

        url = f"/files/{filename}"
        return url, None

    except Exception as e:
        return None, f"Chart generation failed: {e}"


# ── Tool definition ────────────────────────────────────────────────────────────
SQL_TOOL = {
    "name": "execute_sql",
    "description": (
        "Execute a read-only SQL query against the Sealine searates database "
        "(SQL Server). Use this to answer questions with live data. "
        "Only SELECT and WITH (CTE) statements are allowed. "
        f"Results are capped at {MAX_ROWS} rows."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "query": {
                "type": "string",
                "description": "The SQL query to execute (SELECT or WITH only)."
            }
        },
        "required": ["query"]
    }
}


GENERATE_MAP_TOOL = {
    "name": "generate_route_map",
    "description": (
        "Generate an interactive Leaflet.js HTML map for a specific tracking number's container route. "
        "Queries the database for all container events, plots each container's route with coloured "
        "polylines and clickable markers, and returns a /files/ URL you can embed in the chat. "
        "Use this whenever the user asks to 'show the route on a map' or 'visualise the route'."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "track_number": {
                "type": "string",
                "description": "The tracking number (e.g. MEDUAD781707)"
            }
        },
        "required": ["track_number"]
    }
}

CREATE_EXCEL_TOOL = {
    "name": "create_excel",
    "description": (
        "Create an Excel (.xlsx) file from tabular data and return the file path. "
        "Use this when the user asks for an Excel file or wants to email results as a spreadsheet. "
        "After calling this, call send_email to deliver the file."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "title":    {"type": "string",  "description": "Sheet title (e.g. 'IN_TRANSIT Report')"},
            "columns":  {"type": "array",   "items": {"type": "string"}, "description": "Column header names"},
            "rows":     {"type": "array",   "items": {"type": "array"},  "description": "Array of data rows (each row is an array of values matching columns)"},
            "filename": {"type": "string",  "description": "Output filename without extension (e.g. 'intransit_report')"},
        },
        "required": ["title", "columns", "rows", "filename"]
    }
}

SEND_EMAIL_TOOL = {
    "name": "send_email",
    "description": (
        "Send an email via Gmail, optionally with a file attachment (e.g. an Excel file). "
        "Use attachment_path from a previous create_excel call to send results as a spreadsheet. "
        "Default recipient is hangseanshi@gmail.com unless the user specifies otherwise."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "to":              {"type": "string", "description": "Recipient email address"},
            "subject":         {"type": "string", "description": "Email subject line"},
            "body":            {"type": "string", "description": "Email body (plain text or simple HTML)"},
            "attachment_path": {"type": "string", "description": "Full file path to attach (optional — from create_excel)"},
        },
        "required": ["to", "subject", "body"]
    }
}

GENERATE_LOCATION_MAP_TOOL = {
    "name": "generate_location_map",
    "description": (
        "Generate an interactive Leaflet.js HTML map that plots a set of locations as pin markers. "
        "Use this when the user wants to see WHERE multiple containers or shipments currently ARE on a map — "
        "e.g. 'show current locations', 'plot containers on a map', 'show ETA arrivals on a map'. "
        "First query the database to get lat/lng for each container's current location, then call this tool. "
        "Each marker needs lat, lng, label and an optional popup (HTML allowed) and color. "
        "Returns a /files/ URL — include it as a markdown link [View Map](url) so it embeds as an iframe."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "title":    {"type": "string", "description": "Map title shown in the header bar"},
            "markers":  {
                "type": "array",
                "description": "List of marker objects to plot",
                "items": {
                    "type": "object",
                    "properties": {
                        "lat":   {"type": "number", "description": "Latitude"},
                        "lng":   {"type": "number", "description": "Longitude"},
                        "label": {"type": "string", "description": "Short label (e.g. container number)"},
                        "popup": {"type": "string", "description": "Popup HTML content shown on click (optional)"},
                        "color": {"type": "string", "description": "Hex pin colour, e.g. #4e9af1 (optional)"},
                    },
                    "required": ["lat", "lng", "label"]
                }
            },
            "filename": {"type": "string", "description": "Output filename without extension (e.g. 'eta_locations')"},
        },
        "required": ["title", "markers", "filename"]
    }
}

GENERATE_CHART_TOOL = {
    "name": "generate_chart",
    "description": (
        "Generate an interactive Chart.js HTML chart and return a /files/ URL for embedding. "
        "Supports bar, line, pie, doughnut, and horizontalBar chart types. "
        "Use this whenever the user asks to visualise data as a chart or graph. "
        "IMPORTANT: include the returned URL as a markdown link [View Chart](url) so it renders as an embedded iframe."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "title":      {"type": "string", "description": "Chart title shown at the top"},
            "chart_type": {
                "type": "string",
                "enum": ["bar", "line", "pie", "doughnut", "horizontalBar"],
                "description": "Chart.js chart type"
            },
            "labels":     {
                "type": "array",
                "items": {"type": "string"},
                "description": "Category labels along the X-axis (or pie/doughnut slice names)"
            },
            "datasets":   {
                "type": "array",
                "description": "One or more data series",
                "items": {
                    "type": "object",
                    "properties": {
                        "label": {"type": "string", "description": "Series name shown in legend"},
                        "data":  {"type": "array", "items": {"type": "number"}, "description": "Numeric values matching the labels array"}
                    },
                    "required": ["label", "data"]
                }
            },
            "filename":   {"type": "string", "description": "Output filename without extension (e.g. 'intransit_chart')"},
            "x_label":    {"type": "string", "description": "X-axis label (optional, not used for pie/doughnut)"},
            "y_label":    {"type": "string", "description": "Y-axis label (optional, not used for pie/doughnut)"},
        },
        "required": ["title", "chart_type", "labels", "datasets", "filename"]
    }
}


# ── Markdown loader ────────────────────────────────────────────────────────────
def load_md_files(search_root: str) -> tuple[str, list[str]]:
    pattern = os.path.join(search_root, "**", "*.md")
    paths = sorted(glob.glob(pattern, recursive=True))
    if not paths:
        return "", []
    sections, loaded = [], []
    for path in paths:
        try:
            with open(path, encoding="utf-8") as f:
                content = f.read().strip()
            rel = os.path.relpath(path, search_root)
            sections.append(f"## File: {rel}\n\n{content}")
            loaded.append(rel)
        except OSError:
            pass
    return "\n\n---\n\n".join(sections), loaded


# ── Banner ─────────────────────────────────────────────────────────────────────
def make_banner() -> str:
    return f"""
  {c('╔══════════════════════════════════════════╗', CYAN, BOLD)}
  {c('║', CYAN, BOLD)}  {c('Claude for Desktop  (terminal edition)', BOLD)}  {c('║', CYAN, BOLD)}
  {c('╚══════════════════════════════════════════╝', CYAN, BOLD)}

  {c('Commands:', DIM)}
  {c('  /clear', YELLOW)}   — clear conversation history
  {c('  /history', YELLOW)} — show message count & token usage
  {c('  /docs', YELLOW)}    — list loaded Markdown files
  {c('  /system', YELLOW)}  — view/set the system prompt
  {c('  /quit', YELLOW)}    — exit  (or Ctrl-C / Ctrl-D)
  {c('  /help', YELLOW)}    — show this help

  Press {c('Enter', BOLD)} twice on an empty line to submit multi-line input.
"""


# ── Conversation manager ───────────────────────────────────────────────────────
class ClaudeChat:
    def __init__(
        self,
        model: str = "claude-haiku-4-5",
        base_system: str = "You are Claude, a helpful AI assistant made by Anthropic.",
        max_tokens: int = 8192,
        docs_text: str = "",
        docs_files: list[str] | None = None,
        db_enabled: bool = True,
    ):
        self.client = anthropic.Anthropic(
            http_client=httpx.Client(verify=False)
        )
        self.model = model
        self.base_system = base_system
        self.max_tokens = max_tokens
        self.docs_text = docs_text
        self.docs_files = docs_files or []
        self.db_enabled = db_enabled and PYODBC_AVAILABLE
        self.messages: list[dict] = []
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.cache_hits = 0
        self.sql_calls = 0
        self.email_calls = 0
        self.excel_calls = 0

    # ── System blocks ──────────────────────────────────────────────────────────
    def _system_blocks(self) -> list[dict] | str:
        db_note = (
            "\n\nYou have access to the `execute_sql` tool which runs live queries "
            "against the Sealine searates SQL Server database. Use it whenever the "
            "user asks for data, counts, reports, or anything requiring live results."
            if self.db_enabled else ""
        )
        db_note += (
            "\n\nYou also have access to:\n"
            "- `generate_route_map`: generate an interactive HTML map for a SINGLE tracking number's sequential route — "
            "use ONLY when the user asks to see one shipment's route/journey. "
            "Returns a /files/ URL — include as [View Map](url).\n"
            "- `generate_location_map`: plot MULTIPLE containers/shipments as pin markers on an interactive map — "
            "use when the user asks to see WHERE multiple containers currently are, e.g. 'show current locations', "
            "'plot containers on a map', 'show ETA arrivals on a map'. "
            "First query the DB for lat/lng coordinates, then call this tool with the marker list. "
            "Returns a /files/ URL — include as [View Map](url).\n"
            "- `generate_chart`: generate an interactive Chart.js chart (bar, line, pie, doughnut, horizontalBar) — "
            "use ONLY for statistical/numerical visualisations (counts, volumes, trends). "
            "Do NOT use for geographic/location data — use a map tool instead. "
            "Returns a /files/ URL — include as [View Chart](url).\n"
            "- `create_excel`: build a formatted Excel (.xlsx) file from tabular data\n"
            "- `send_email`: send an email via Gmail with an optional Excel attachment\n"
            "When the user asks to email results or create an Excel file, use these tools. "
            "Default email recipient is hangseanshi@gmail.com unless specified otherwise."
        )
        base = self.base_system + db_note

        if not self.docs_text:
            return base

        return [
            {"type": "text", "text": base},
            {
                "type": "text",
                "text": (
                    "# Sealine-Database Reference Documents\n\n"
                    "The following Markdown files have been loaded from the repository. "
                    "Use them as your primary reference for schema, relationships, "
                    "connection details, and saved reports.\n\n"
                    + self.docs_text
                ),
                "cache_control": {"type": "ephemeral"},
            },
        ]

    # ── Tool executor ──────────────────────────────────────────────────────────
    def _execute_tool(self, name: str, tool_input: dict) -> str:
        if name == "execute_sql":
            query = tool_input.get("query", "")
            print(c(f"\n  [SQL] ", YELLOW, BOLD), end="")
            short = query.replace("\n", " ").strip()
            print(c(textwrap.shorten(short, width=100, placeholder="…"), DIM))
            result = run_sql(query)
            self.sql_calls += 1
            return result

        if name == "generate_route_map":
            tn = tool_input.get("track_number", "")
            print(c(f"\n  [MAP] Generating route map for {tn} …", CYAN, BOLD))
            url, err = generate_route_map(tn)
            if err:
                return f"MAP ERROR: {err}"
            print(c(f"  [MAP] Saved → {url}", DIM))
            return f"Route map generated: {url}"

        if name == "create_excel":
            print(c(f"\n  [EXCEL] Creating {tool_input.get('filename','report')}.xlsx …", CYAN, BOLD))
            filepath, err = create_excel_file(
                title    = tool_input.get("title", "Report"),
                columns  = tool_input.get("columns", []),
                rows     = tool_input.get("rows", []),
                filename = tool_input.get("filename", "report"),
            )
            self.excel_calls += 1
            if err:
                return f"EXCEL ERROR: {err}"
            fname = os.path.basename(filepath)
            print(c(f"  [EXCEL] Saved → {filepath}", DIM))
            return f"Excel file created: /files/{fname}"

        if name == "send_email":
            to      = tool_input.get("to", "hangseanshi@gmail.com")
            subject = tool_input.get("subject", "(no subject)")
            body    = tool_input.get("body", "")
            att     = tool_input.get("attachment_path")
            print(c(f"\n  [EMAIL] Sending to {to}: {subject}", MAGENTA, BOLD))
            result = send_email_with_attachment(to, subject, body, att)
            self.email_calls += 1
            return result

        if name == "generate_chart":
            print(c(f"\n  [CHART] Generating {tool_input.get('chart_type','bar')} chart: {tool_input.get('title','')} …", CYAN, BOLD))
            url, err = generate_chart(
                title      = tool_input.get("title", "Chart"),
                chart_type = tool_input.get("chart_type", "bar"),
                labels     = tool_input.get("labels", []),
                datasets   = tool_input.get("datasets", []),
                filename   = tool_input.get("filename", "chart"),
                x_label    = tool_input.get("x_label", ""),
                y_label    = tool_input.get("y_label", ""),
            )
            if err:
                return f"CHART ERROR: {err}"
            print(c(f"  [CHART] Saved → {url}", DIM))
            return f"Chart generated: {url}"

        if name == "generate_location_map":
            print(c(f"\n  [MAP] Generating location map: {tool_input.get('title','')} …", CYAN, BOLD))
            url, err = generate_location_map(
                title    = tool_input.get("title", "Locations"),
                markers  = tool_input.get("markers", []),
                filename = tool_input.get("filename", "locations"),
            )
            if err:
                return f"MAP ERROR: {err}"
            print(c(f"  [MAP] Saved → {url}", DIM))
            return f"Location map generated: {url}"

        return f"Unknown tool: {name}"

    # ── Send (API — returns response text) ──────────────────────────────────
    def send_api(self, user_text: str) -> str:
        """Non-streaming send that returns the full response text. Used by the REST API."""
        self.messages.append({"role": "user", "content": user_text})
        tools = ([SQL_TOOL] if self.db_enabled else []) + [GENERATE_MAP_TOOL, GENERATE_LOCATION_MAP_TOOL, CREATE_EXCEL_TOOL, SEND_EMAIL_TOOL, GENERATE_CHART_TOOL]

        while True:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=self.max_tokens,
                system=self._system_blocks(),
                tools=tools,
                messages=self.messages,
            )

            self.total_input_tokens  += response.usage.input_tokens
            self.total_output_tokens += response.usage.output_tokens
            cache_read = getattr(response.usage, "cache_read_input_tokens", 0) or 0
            if cache_read:
                self.cache_hits += 1

            self.messages.append({"role": "assistant", "content": response.content})

            if response.stop_reason == "tool_use":
                tool_results = []
                for block in response.content:
                    if block.type == "tool_use":
                        result_text = self._execute_tool_silent(block.name, block.input)
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result_text,
                        })
                self.messages.append({"role": "user", "content": tool_results})
                continue

            # Extract text from response
            parts = []
            for block in response.content:
                if block.type == "text":
                    parts.append(block.text)
            return "\n".join(parts)

    def _execute_tool_silent(self, name: str, tool_input: dict) -> str:
        """Execute tool without terminal output (for API use)."""
        if name == "execute_sql":
            query = tool_input.get("query", "")
            result = run_sql(query)
            self.sql_calls += 1
            return result

        if name == "generate_route_map":
            tn = tool_input.get("track_number", "")
            url, err = generate_route_map(tn)
            if err:
                return f"MAP ERROR: {err}"
            return f"Route map generated: {url}"

        if name == "create_excel":
            filepath, err = create_excel_file(
                title    = tool_input.get("title", "Report"),
                columns  = tool_input.get("columns", []),
                rows     = tool_input.get("rows", []),
                filename = tool_input.get("filename", "report"),
            )
            self.excel_calls += 1
            if err:
                return f"EXCEL ERROR: {err}"
            fname = os.path.basename(filepath)
            return f"Excel file created: /files/{fname}"

        if name == "send_email":
            to      = tool_input.get("to", "hangseanshi@gmail.com")
            subject = tool_input.get("subject", "(no subject)")
            body    = tool_input.get("body", "")
            att     = tool_input.get("attachment_path")
            result  = send_email_with_attachment(to, subject, body, att)
            self.email_calls += 1
            return result

        if name == "generate_chart":
            url, err = generate_chart(
                title      = tool_input.get("title", "Chart"),
                chart_type = tool_input.get("chart_type", "bar"),
                labels     = tool_input.get("labels", []),
                datasets   = tool_input.get("datasets", []),
                filename   = tool_input.get("filename", "chart"),
                x_label    = tool_input.get("x_label", ""),
                y_label    = tool_input.get("y_label", ""),
            )
            if err:
                return f"CHART ERROR: {err}"
            return f"Chart generated: {url}"

        if name == "generate_location_map":
            url, err = generate_location_map(
                title    = tool_input.get("title", "Locations"),
                markers  = tool_input.get("markers", []),
                filename = tool_input.get("filename", "locations"),
            )
            if err:
                return f"MAP ERROR: {err}"
            return f"Location map generated: {url}"

        return f"Unknown tool: {name}"

    # ── Send (agentic loop with streaming) ────────────────────────────────────
    def send(self, user_text: str) -> None:
        self.messages.append({"role": "user", "content": user_text})
        print(f"\n{c('Claude', CYAN, BOLD)}  ", end="", flush=True)

        tools = ([SQL_TOOL] if self.db_enabled else []) + [GENERATE_MAP_TOOL, GENERATE_LOCATION_MAP_TOOL, CREATE_EXCEL_TOOL, SEND_EMAIL_TOOL, GENERATE_CHART_TOOL]

        while True:
            collected: list[str] = []

            # Stream one API call
            try:
                with self.client.messages.stream(
                    model=self.model,
                    max_tokens=self.max_tokens,
                    system=self._system_blocks(),
                    tools=tools,
                    messages=self.messages,
                ) as stream:
                    for event in stream:
                        if event.type == "content_block_start":
                            if event.content_block.type == "thinking":
                                print(c("\n[thinking…]", DIM, MAGENTA), flush=True)
                            elif event.content_block.type == "tool_use":
                                # tool name shown after we get the full input
                                pass
                        elif event.type == "content_block_delta":
                            if event.delta.type == "text_delta":
                                print(event.delta.text, end="", flush=True)
                                collected.append(event.delta.text)
                    final = stream.get_final_message()

            except anthropic.BadRequestError:
                # Haiku/older model — no thinking support, retry without it
                with self.client.messages.stream(
                    model=self.model,
                    max_tokens=self.max_tokens,
                    system=self._system_blocks(),
                    tools=tools,
                    messages=self.messages,
                ) as stream:
                    for text in stream.text_stream:
                        print(text, end="", flush=True)
                        collected.append(text)
                    final = stream.get_final_message()

            self.total_input_tokens  += final.usage.input_tokens
            self.total_output_tokens += final.usage.output_tokens
            cache_read = getattr(final.usage, "cache_read_input_tokens", 0) or 0
            if cache_read:
                self.cache_hits += 1

            # Append assistant turn (full content list, preserves tool_use blocks)
            self.messages.append({"role": "assistant", "content": final.content})

            # ── Check if Claude wants to use a tool ───────────────────────────
            if final.stop_reason == "tool_use":
                tool_results = []
                for block in final.content:
                    if block.type == "tool_use":
                        result_text = self._execute_tool(block.name, block.input)
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result_text,
                        })
                # Feed results back and loop
                self.messages.append({"role": "user", "content": tool_results})
                print(f"\n{c('Claude', CYAN, BOLD)}  ", end="", flush=True)
                continue

            # ── end_turn: done ────────────────────────────────────────────────
            break

        print("\n")

    # ── Slash commands ─────────────────────────────────────────────────────────
    def cmd_clear(self) -> None:
        self.messages.clear()
        print(c("  Conversation cleared.\n", DIM))

    def cmd_history(self) -> None:
        turns = sum(1 for m in self.messages if m["role"] == "user"
                    and not (isinstance(m["content"], list)
                             and m["content"] and m["content"][0].get("type") == "tool_result"))
        print(
            f"\n  {c('Turns:', BOLD)} {turns}  |  "
            f"{c('Input tokens:', BOLD)} {self.total_input_tokens:,}  |  "
            f"{c('Output tokens:', BOLD)} {self.total_output_tokens:,}  |  "
            f"{c('Cache hits:', BOLD)} {self.cache_hits}  |  "
            f"{c('SQL calls:', BOLD)} {self.sql_calls}\n"
        )

    def cmd_docs(self) -> None:
        if not self.docs_files:
            print(c("  No Markdown files loaded.\n", DIM))
            return
        print(f"\n  {c('Loaded Markdown files:', BOLD)}")
        for f in self.docs_files:
            print(f"    {c('•', CYAN)} {f}")
        print()

    def cmd_system(self, rest: str) -> None:
        if rest.strip():
            self.base_system = rest.strip()
            self.messages.clear()
            print(c("  System prompt updated (history cleared).\n", DIM))
        else:
            print(f"\n  {c('System prompt:', BOLD)}\n  {self.base_system}\n")
            if self.docs_files:
                print(c(f"  + {len(self.docs_files)} Markdown file(s) loaded as cached context\n", DIM))

    def cmd_help(self) -> None:
        print(make_banner())


# ── Input helpers ──────────────────────────────────────────────────────────────
def read_input(prompt: str) -> str:
    try:
        first = input(prompt)
    except (EOFError, KeyboardInterrupt):
        raise
    if first == "":
        return ""
    lines = [first]
    while True:
        try:
            line = input()
        except (EOFError, KeyboardInterrupt):
            break
        if line == "":
            break
        lines.append(line)
    return "\n".join(lines).strip()


# ── Main REPL ──────────────────────────────────────────────────────────────────
def main() -> None:
    script_dir = os.path.dirname(os.path.abspath(__file__))

    parser = argparse.ArgumentParser(
        description="Terminal chat interface for Claude with live Sealine DB access"
    )
    parser.add_argument("--model", default="claude-haiku-4-5",
                        help="Model ID (default: claude-haiku-4-5)")
    parser.add_argument("--system",
                        default="You are Claude, a helpful AI assistant and data analyst "
                                "for the Sealine shipping database. You have been given "
                                "the database schema and reference documents as context.",
                        help="Base system prompt")
    parser.add_argument("--max-tokens", type=int, default=8192,
                        help="Max output tokens per response (default: 8192)")
    parser.add_argument("--docs-dir", default=script_dir,
                        help=f"Root directory to search for .md files (default: {script_dir})")
    parser.add_argument("--no-docs", action="store_true",
                        help="Skip loading Markdown files")
    parser.add_argument("--no-db", action="store_true",
                        help="Disable live SQL database tool")
    args = parser.parse_args()

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print(c("Error: ANTHROPIC_API_KEY environment variable is not set.", RED, BOLD))
        sys.exit(1)

    docs_text, docs_files = "", []
    if not args.no_docs:
        docs_text, docs_files = load_md_files(args.docs_dir)

    db_enabled = not args.no_db
    if db_enabled and not PYODBC_AVAILABLE:
        print(c("  Warning: pyodbc not installed — DB tool disabled. Run: pip install pyodbc", YELLOW))
        db_enabled = False

    chat = ClaudeChat(
        model=args.model,
        base_system=args.system,
        max_tokens=args.max_tokens,
        docs_text=docs_text,
        docs_files=docs_files,
        db_enabled=db_enabled,
    )

    print(make_banner())
    print(c(f"  Model    : {chat.model}", DIM))
    print(c(f"  System   : {chat.base_system[:80]}{'…' if len(chat.base_system) > 80 else ''}", DIM))

    if docs_files:
        print(c(f"  Docs     : {len(docs_files)} Markdown file(s) loaded & cached", GREEN))
        for f in docs_files:
            print(c(f"             • {f}", DIM))
    else:
        print(c("  Docs     : none", DIM))

    if chat.db_enabled:
        print(c("  Database : connected — live SQL queries enabled", GREEN))
        print(c("             ushou102-exap1 / searates", DIM))
    else:
        print(c("  Database : disabled (use --no-db to suppress, or install pyodbc)", DIM))

    print()

    user_prompt = f"{c('You', GREEN, BOLD)}  "

    while True:
        try:
            text = read_input(user_prompt)
        except (EOFError, KeyboardInterrupt):
            print(c("\n\n  Goodbye!\n", DIM))
            break

        if not text:
            continue

        if text.startswith("/"):
            cmd, _, rest = text.partition(" ")
            cmd = cmd.lower()
            if cmd in ("/quit", "/exit", "/q"):
                print(c("\n  Goodbye!\n", DIM))
                break
            elif cmd == "/clear":
                chat.cmd_clear()
            elif cmd == "/history":
                chat.cmd_history()
            elif cmd == "/docs":
                chat.cmd_docs()
            elif cmd == "/system":
                chat.cmd_system(rest)
            elif cmd == "/help":
                chat.cmd_help()
            else:
                print(c(f"  Unknown command '{cmd}'. Type /help for help.\n", RED))
            continue

        try:
            chat.send(text)
        except anthropic.AuthenticationError:
            print(c("\n  Error: Invalid API key. Check ANTHROPIC_API_KEY.\n", RED))
        except anthropic.RateLimitError:
            print(c("\n  Error: Rate limited. Please wait and try again.\n", RED))
        except anthropic.APIConnectionError:
            print(c("\n  Error: Network error. Check your internet connection.\n", RED))
        except anthropic.APIStatusError as e:
            print(c(f"\n  API error {e.status_code}: {e.message}\n", RED))


if __name__ == "__main__":
    main()
